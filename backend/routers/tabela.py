"""
routers/tabela.py — Endpoint para visualização tabular dos dados de lift.

Retorna os registros flat (não em árvore) com os mesmos parâmetros
de filtro do dendrograma. Usado pela aba de tabela do frontend.
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO, StringIO
import csv

from backend.database import get_db
from backend.models.lift import LiftResultado, Onda

router = APIRouter(prefix="/api", tags=["tabela"])


@router.get("/tabela")
def get_tabela(
    onda: str = Query(..., description="Código da onda"),
    assunto: str = Query(..., description="ASSUNTO_COLUNA"),
    pergunta: str = Query(..., description="PERGUNTA_COLUNA"),
    categoria_coluna: str | None = Query(None),
    direcao: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """
    Retorna os registros flat para exibição em tabela.
    Mesmos filtros do /api/tree — mas retorna linhas, não árvore.
    """
    onda_obj = db.query(Onda).filter_by(codigo=onda).first()
    if not onda_obj:
        raise HTTPException(status_code=404, detail=f"Onda '{onda}' não encontrada.")

    query = db.query(LiftResultado).filter(
        LiftResultado.onda_id == onda_obj.id,
        LiftResultado.assunto_coluna == assunto,
        LiftResultado.pergunta_coluna == pergunta,
    )

    if categoria_coluna:
        query = query.filter(LiftResultado.categoria_coluna == categoria_coluna)

    if direcao:
        query = query.filter(LiftResultado.direcao == direcao)

    query = query.order_by(LiftResultado.rank_global)

    rows = query.all()

    if not rows:
        raise HTTPException(status_code=404, detail="Nenhum resultado encontrado.")

    return [_serialize_row(r) for r in rows]


@router.get("/tabela/download/csv")
def download_csv(
    onda: str = Query(...),
    assunto: str = Query(...),
    pergunta: str = Query(...),
    categoria_coluna: str | None = Query(None),
    direcao: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Download da tabela filtrada em formato CSV."""
    rows = _get_rows(db, onda, assunto, pergunta, categoria_coluna, direcao)

    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=_COLUMNS.keys())
    writer.writeheader()
    for row in rows:
        writer.writerow(_serialize_row(row))

    output.seek(0)
    filename = f"nexas_{onda}_{assunto[:20]}.csv".replace(" ", "_").replace("/", "-")

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/tabela/download/xlsx")
def download_xlsx(
    onda: str = Query(...),
    assunto: str = Query(...),
    pergunta: str = Query(...),
    categoria_coluna: str | None = Query(None),
    direcao: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Download da tabela filtrada em formato Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado.")

    rows = _get_rows(db, onda, assunto, pergunta, categoria_coluna, direcao)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NEXAS"

    # Header
    headers = list(_COLUMNS.values())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(rows, 2):
        data = _serialize_row(row)
        for col_idx, key in enumerate(_COLUMNS.keys(), 1):
            ws.cell(row=row_idx, column=col_idx, value=data.get(key))

    # Ajuste de largura
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"nexas_{onda}_{assunto[:20]}.xlsx".replace(" ", "_").replace("/", "-")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================
# Helpers internos
# ============================================

# Mapeamento chave_interna → label da coluna na tabela
_COLUMNS = {
    "categoria_coluna":   "CATEGORIA_COLUNA",
    "assunto_linha":      "ASSUNTO_LINHA",
    "pergunta_linha":     "PERGUNTA_LINHA",
    "categoria_linha":    "CATEGORIA_LINHA",
    "lift":               "LIFT",
    "score_absoluto":     "SCORE ABSOLUTO",
    "direcao":            "DIREÇÃO",
    "categoria_direcao":  "CATEGORIA DIREÇÃO",
    "rank_global":        "RANK GLOBAL",
    "percentil_relevancia": "PERCENTIL",
    "ranking_final":      "RANKING FINAL",
    "per_relativo":       "% RELATIVO",
}


def _serialize_row(r: LiftResultado) -> dict:
    return {
        "categoria_coluna":      r.categoria_coluna,
        "assunto_linha":         r.assunto_linha,
        "pergunta_linha":        r.pergunta_linha,
        "categoria_linha":       r.categoria_linha,
        "lift":                  float(r.lift) if r.lift else None,
        "score_absoluto":        float(r.score_absoluto) if r.score_absoluto else None,
        "direcao":               r.direcao,
        "categoria_direcao":     r.categoria_direcao,
        "rank_global":           r.rank_global,
        "percentil_relevancia":  float(r.percentil_relevancia) if r.percentil_relevancia else None,
        "ranking_final":         r.ranking_final,
        "per_relativo":          float(r.per_relativo) if r.per_relativo is not None else None,
    }


def _get_rows(db, onda, assunto, pergunta, categoria_coluna, direcao):
    onda_obj = db.query(Onda).filter_by(codigo=onda).first()
    if not onda_obj:
        raise HTTPException(status_code=404, detail=f"Onda '{onda}' não encontrada.")

    query = db.query(LiftResultado).filter(
        LiftResultado.onda_id == onda_obj.id,
        LiftResultado.assunto_coluna == assunto,
        LiftResultado.pergunta_coluna == pergunta,
    )
    if categoria_coluna:
        query = query.filter(LiftResultado.categoria_coluna == categoria_coluna)
    if direcao:
        query = query.filter(LiftResultado.direcao == direcao)

    return query.order_by(LiftResultado.rank_global).all()
